import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

def generate_excel_report(notas: list, output_folder: str, filename: str = 'Relatorio_NFe.xlsx') -> str:
    outdir = Path(output_folder)
    outdir.mkdir(parents=True, exist_ok=True)
    outpath = outdir / filename

    # Resumo por nota
    resumo_rows = []
    detalhes_rows = []
    for n in notas:
        resumo_rows.append({
            'Arquivo': n.get('arquivo'),
            'Chave NFe': n.get('chave'),
            'Emitente': n.get('emit_xnome'),
            'CNPJ Emitente': n.get('emit_cnpj'),
            'Destinatário': n.get('dest_xnome'),
            'CNPJ Destinatário': n.get('dest_cnpj'),
            'Data Emissão': n.get('data_emissao'),
            'Valor Total': n.get('valor_total'),
            'Total Itens': n.get('total_itens'),
        })
        for p in n['produtos']:
            detalhes_rows.append({
                'Arquivo': n.get('arquivo'),
                'Chave NFe': n.get('chave'),
                'Descrição': p['descricao'],
                'Quantidade': p['quantidade'],
                'Valor Unitário': p['valor_unitario'],
                'Valor Total Item': p['valor_total_item'],
            })

    # Cria Excel
    resumo_df = pd.DataFrame(resumo_rows)
    detalhes_df = pd.DataFrame(detalhes_rows)

    with pd.ExcelWriter(outpath, engine='openpyxl') as writer:
        resumo_df.to_excel(writer, sheet_name='Resumo', index=False)
        detalhes_df.to_excel(writer, sheet_name='Detalhes', index=False)

    # Formatação
    wb = load_workbook(outpath)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Cabeçalho em negrito e fundo azul claro
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        # Largura automática colunas
        for i, col in enumerate(ws.columns, 1):
            max_length = max([len(str(c.value)) if c.value else 0 for c in col])
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2
    wb.save(outpath)
    return str(outpath)
